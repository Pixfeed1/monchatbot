"""
Document Processor - Extraction de contenu depuis différents formats
Supporte: PDF, Word (DOC/DOCX), Excel (XLS/XLSX), TXT
"""

import logging
import os
from typing import Optional, Tuple

logger = logging.getLogger(__name__)


class DocumentProcessor:
    """Processeur de documents pour extraction de contenu"""

    def __init__(self):
        self.supported_extensions = ['.pdf', '.doc', '.docx', '.txt', '.xls', '.xlsx']

    def is_supported(self, filename: str) -> bool:
        """Vérifie si le format de fichier est supporté."""
        ext = os.path.splitext(filename)[1].lower()
        return ext in self.supported_extensions

    def extract_content(self, file_path: str) -> Tuple[str, Optional[str]]:
        """
        Extrait le contenu d'un document.

        Args:
            file_path: Chemin vers le fichier

        Returns:
            Tuple (content, summary) - Contenu extrait et résumé optionnel
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Fichier introuvable: {file_path}")

        ext = os.path.splitext(file_path)[1].lower()

        try:
            if ext == '.txt':
                return self._extract_txt(file_path)
            elif ext == '.pdf':
                return self._extract_pdf(file_path)
            elif ext in ['.doc', '.docx']:
                return self._extract_word(file_path)
            elif ext in ['.xls', '.xlsx']:
                return self._extract_excel(file_path)
            else:
                logger.warning(f"Format non supporté: {ext}")
                return "", None

        except Exception as e:
            logger.error(f"Erreur extraction contenu de {file_path}: {e}")
            return "", None

    def _extract_txt(self, file_path: str) -> Tuple[str, Optional[str]]:
        """Extrait le contenu d'un fichier texte."""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()

            summary = self._generate_summary(content)
            return content, summary

        except UnicodeDecodeError:
            # Essayer avec d'autres encodages
            try:
                with open(file_path, 'r', encoding='latin-1') as f:
                    content = f.read()
                summary = self._generate_summary(content)
                return content, summary
            except Exception as e:
                logger.error(f"Erreur lecture TXT: {e}")
                return "", None

    def _extract_pdf(self, file_path: str) -> Tuple[str, Optional[str]]:
        """Extrait le contenu d'un fichier PDF."""
        try:
            import PyPDF2

            content = ""
            with open(file_path, 'rb') as f:
                pdf_reader = PyPDF2.PdfReader(f)
                num_pages = len(pdf_reader.pages)

                for page_num in range(num_pages):
                    page = pdf_reader.pages[page_num]
                    content += page.extract_text() + "\n"

            summary = self._generate_summary(content)
            return content.strip(), summary

        except ImportError:
            logger.warning("PyPDF2 non installé. Installation requise: pip install PyPDF2")
            return "", None
        except Exception as e:
            logger.error(f"Erreur extraction PDF: {e}")
            return "", None

    def _extract_word(self, file_path: str) -> Tuple[str, Optional[str]]:
        """Extrait le contenu d'un fichier Word (DOCX)."""
        try:
            from docx import Document

            doc = Document(file_path)
            content = "\n".join([paragraph.text for paragraph in doc.paragraphs])

            summary = self._generate_summary(content)
            return content, summary

        except ImportError:
            logger.warning("python-docx non installé. Installation requise: pip install python-docx")
            return "", None
        except Exception as e:
            logger.error(f"Erreur extraction Word: {e}")
            return "", None

    def _extract_excel(self, file_path: str) -> Tuple[str, Optional[str]]:
        """Extrait le contenu d'un fichier Excel."""
        try:
            import openpyxl

            workbook = openpyxl.load_workbook(file_path, read_only=True)
            content = []

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                content.append(f"=== Feuille: {sheet_name} ===\n")

                for row in sheet.iter_rows(values_only=True):
                    row_text = "\t".join([str(cell) if cell is not None else "" for cell in row])
                    if row_text.strip():
                        content.append(row_text)

            full_content = "\n".join(content)
            summary = self._generate_summary(full_content)
            return full_content, summary

        except ImportError:
            logger.warning("openpyxl non installé. Installation requise: pip install openpyxl")
            return "", None
        except Exception as e:
            logger.error(f"Erreur extraction Excel: {e}")
            return "", None

    def _generate_summary(self, content: str, max_length: int = 500) -> Optional[str]:
        """
        Génère un résumé simple du contenu.

        Args:
            content: Contenu à résumer
            max_length: Longueur maximale du résumé

        Returns:
            Résumé du contenu
        """
        if not content:
            return None

        # Résumé simple: premiers caractères + statistiques
        content = content.strip()
        lines = content.split('\n')
        words = content.split()

        if len(content) <= max_length:
            summary = content
        else:
            # Prendre les premiers mots jusqu'à max_length
            summary = content[:max_length] + "..."

        # Ajouter des statistiques
        stats = f"\n\nStatistiques: {len(lines)} lignes, {len(words)} mots, {len(content)} caractères"
        summary += stats

        return summary

    def process_and_update_document(self, document_model, file_path: str) -> bool:
        """
        Traite un document et met à jour le modèle avec le contenu extrait.

        Args:
            document_model: Instance du modèle Document
            file_path: Chemin vers le fichier

        Returns:
            True si succès, False sinon
        """
        try:
            content, summary = self.extract_content(file_path)

            if content:
                document_model.content = content
                document_model.summary = summary
                document_model.status = 'processed'
                logger.info(f"Document {document_model.filename} traité avec succès")
                return True
            else:
                document_model.status = 'error'
                logger.warning(f"Impossible d'extraire le contenu de {document_model.filename}")
                return False

        except Exception as e:
            logger.error(f"Erreur traitement document: {e}")
            document_model.status = 'error'
            return False


# Instance globale
document_processor = DocumentProcessor()
